#!/usr/bin/env python3
"""
解析工作簿13.xlsx中的中国传统村落数据
提取结构化信息(省份、城市、区县、乡镇、村名)并导出为JSON
"""
import openpyxl
import re
import json
from collections import defaultdict

def parse_villages(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Sheet1']

    all_villages = []  # 每个村落一条记录
    current_batch = None
    current_province = None
    current_batch_num = None

    # 批次匹配
    batch_pat = re.compile(r'第([一二三四五六七八九十\d]+)批')

    # 省份匹配 - 支持多种格式
    # "北京市(9个)", "河北省（32个）", "一、北京市(9个)", "1.北京市(9个)"
    province_pat = re.compile(
        r'^[一二三四五六七八九十\d]+[、.、\s]*'  # 前缀编号(可选)
        r'(.+?)'                                   # 省份名
        r'[（(](\d+)个[）)]'                        # 数量
    )

    # 简单的省份匹配(无编号前缀)
    province_simple_pat = re.compile(r'^(.+?)[（(](\d+)个[）)]$')

    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if not val:
            continue
        val = str(val).strip()
        # 规范化空格
        val = val.replace('　', '').replace('\xa0', ' ').replace('\t', '')
        val = re.sub(r'\s+', '', val)

        # 检查批次标题
        batch_m = batch_pat.search(val)
        if batch_m and ('批' in val) and ('列入' in val or '公示' in val):
            batch_num_str = batch_m.group(1)
            # 转换中文数字
            cn_nums = {'一':'1','二':'2','三':'3','四':'4','五':'5','六':'6','七':'7','八':'8','九':'9','十':'10'}
            current_batch_num = cn_nums.get(batch_num_str, batch_num_str)
            current_batch = val
            current_province = None
            continue

        # 检查省份标题
        m = province_pat.match(val)
        if not m:
            m = province_simple_pat.match(val)

        if m and current_batch:
            prov_name = m.group(1).strip()
            prov_count = int(m.group(2))
            # 修正已知数据问题
            if prov_name == '川省':
                prov_name = '四川省'
            current_province = {
                'name': prov_name,
                'count': prov_count,
                'batch': current_batch,
                'batch_num': current_batch_num
            }
            continue

        # 村落条目
        if current_province is not None and current_batch:
            # 跳过非村落的内容
            if any(kw in val for kw in ['列入', '公示', '传统村落', '附件']):
                continue
            if len(val) < 3:
                continue

            # 解析行政区划层级
            parsed = parse_village_address(val)
            parsed['province'] = current_province['name']
            parsed['batch'] = current_batch
            parsed['batch_num'] = current_batch_num
            all_villages.append(parsed)

    return all_villages


def parse_village_address(address):
    """解析村落地址，提取市/区县/乡镇/村名"""
    result = {
        'full_address': address,
        'city': '',
        'district': '',
        'county': '',
        'town': '',
        'village': ''
    }

    # 常见行政区划后缀
    city_suffixes = r'(?:市|地区|自治州|州|盟|林区|管理区)'
    county_suffixes = r'(?:区|县|市|自治县|自治旗|旗)'
    town_suffixes = r'(?:镇|乡|街道|民族乡|苏木)'
    village_suffixes = r'(?:村|庄|寨|屯|社区|居委会|组|堡|湾|寨|坝|街|巷|弄)'

    # 提取地名层级
    # 模式: ...市...县...镇...村
    # 也可能是: ...区...镇...村 或 ...县...乡...村

    addr = address

    # 方法: 按后缀分割
    parts = []
    remaining = addr

    # 尝试匹配各级
    # 市级别
    city_m = re.match(r'^(.+?' + city_suffixes + r')', remaining)
    if city_m:
        result['city'] = city_m.group(1)
        remaining = remaining[len(city_m.group(1)):]
    else:
        # 可能以区/县开头(直辖市情况)
        pass

    # 区县级
    county_m = re.match(r'^(.+?' + county_suffixes + r')', remaining)
    if county_m:
        result['district'] = county_m.group(1)
        remaining = remaining[len(county_m.group(1)):]

    # 乡镇级
    town_m = re.match(r'^(.+?' + town_suffixes + r')', remaining)
    if town_m:
        result['town'] = town_m.group(1)
        remaining = remaining[len(town_m.group(1)):]

    # 剩下的就是村名
    if remaining:
        result['village'] = remaining
        # 尝试从村名中分离出更具体的信息
        # 最后的"村"/"庄"/"寨"等之前的内容作为村名主体

    return result


def build_stats(villages):
    """构建统计数据"""
    stats = {
        'total_villages': len(villages),
        'by_province': defaultdict(int),
        'by_batch': defaultdict(int),
        'by_city': defaultdict(int),
        'batches': []
    }

    for v in villages:
        stats['by_province'][v['province']] += 1
        stats['by_batch'][v['batch_num']] += 1
        if v['city']:
            stats['by_city'][v['city']] += 1

    return stats


if __name__ == '__main__':
    filepath = '工作簿13.xlsx'
    print(f"正在解析 {filepath}...")

    villages = parse_villages(filepath)
    print(f"共解析 {len(villages)} 个村落")

    # 保存完整数据
    output = {
        'villages': villages,
        'total': len(villages)
    }

    with open('villages_data.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"数据已保存至 villages_data.json")

    # 打印统计
    stats = build_stats(villages)
    print(f"\n=== 统计 ===")
    print(f"村落总数: {stats['total_villages']}")

    print(f"\n各省村落数量 (Top 20):")
    for prov, cnt in sorted(stats['by_province'].items(), key=lambda x: -x[1])[:20]:
        print(f"  {prov}: {cnt}")

    print(f"\n各批次村落数量:")
    for batch_num in sorted(stats['by_batch'].keys(), key=lambda x: int(x)):
        print(f"  第{batch_num}批: {stats['by_batch'][batch_num]}")

    print(f"\n城市数量: {len(stats['by_city'])}")

    # 打印示例
    print(f"\n=== 解析示例 ===")
    for v in villages[:5]:
        print(f"  {v['full_address']}")
        print(f"    -> 省: {v['province']}, 市: {v['city']}, 区县: {v['district']}, 乡镇: {v['town']}, 村: {v['village']}")
